import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws = wb.active
ws.title = "nsls packages data"
ws.append(
    [
        "Repo",
        "condaforge version",
        "nsls2forge version",
        "missing in conda-forge",
        "PR URLS's",
        "feedstock URLS",
        "anaconda.org URLS",
        "Notes",
    ]
)
redFill = PatternFill(
    start_color="00FF8080", end_color="00FF8080", fill_type="solid"
)  # header for chart, for the color 00FF8080: https://openpyxl.readthedocs.io/en/stable/styles.html#indexed-colours

data = []

with open("xldata.yml", "r") as package_data, open(
    "names.txt", "r"
) as repos:  # yaml file with package data & names.txt file for i in loop
    readinfo = yaml.load(package_data, Loader=yaml.FullLoader)
    data = repos.readlines()

for i in range(2, len(data) + 2):
    cellref = ws.cell(row=i, column=columns['package_name'])
    cellref.value = readinfo[f"package{i-2}"]["package_name"]

    cellref = ws.cell(row=i, column=columns['conda_version'])
    cellref.value = readinfo[f"package{i-2}"]["conda_version"]

    cellref = ws.cell(row=i, column=columns['nsls2_version'])
    cellref.value = readinfo[f"package{i-2}"]["nsls2_version"]

    if readinfo[f"package{i-2}"]["feedstock_URL"] == "no feedstock exists":
        cellref = ws.cell(row=i, column=columns['feedstock_URL'])
        cellref.value = readinfo[f"package{i-2}"]["feedstock_URL"]
    else:
        cellref = ws.cell(row=i, column=columns['feedstock_URL']).hyperlink = readinfo[f"package{i-2}"][
            "feedstock_URL"
        ]

    if readinfo[f"package{i-2}"]["anaconda_URL"] == "no anaconda package exists":
        cellref = ws.cell(row=i, column=columns['anaconda_URL'])
        cellref.value = readinfo[f"package{i-2}"]["anaconda_URL"]
    else:
        cellref = ws.cell(row=i, column=columns['anaconda_URL']).hyperlink = readinfo[f"package{i-2}"][
            "anaconda_URL"
        ]

for cell in ws["B"]:  # highlights the values not found
    if "not found" in cell.value:
        for cell in ws[cell.row]:
            cell.fill = redFill
columns = {
    'package_name': 1,
    'conda_version': 2,
    'nsls2_version': 3,
    'feedstock_URL': 6,
    'anaconda_URL': 7,
}

wb.save("nsls_package_list.xlsx")
