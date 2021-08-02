#!/usr/bin/env python
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.styles import colors


wb = Workbook()
ws = wb.active
ws.title = "nsls packages data"
ws.append(["Repo", "condaforge version", "nsls2forge version", "missing in conda-forge", "PR URLS's", "feedstock URLS", "anaconda.org URLS", "Notes"])
redFill = PatternFill(start_color='00FF8080', end_color='00FF8080', fill_type='solid')  # header for chart


data = []
condafdata = []
nslsfdata = []
githubURLS = []
anacondaURLS = []

with open('names.txt', 'r') as repos, open("conda-forge-vrs.txt", 'r') as condafile, open("nsls2forge-vrs.txt" , 'r') as nsls2f:
    data = repos.readlines()
    condafdata = condafile.readlines()
    nslsfdata = nsls2f.readlines()

with open("feedstock-URLS.txt", 'r') as gitfile, open("anaconda-URLS.txt", 'r') as anacondaf:
    githubURLS = gitfile.readlines()
    anacondaURLS = anacondaf.readlines()

for i in range(2, len(data)+2):
    cellref = ws.cell(row=i, column=1)
    cellref.value = data[i-2]

    cellref = ws.cell(row=i, column=2)
    cellref.value = condafdata[i-2]

    cellref = ws.cell(row=i, column=3)
    cellref.value = nslsfdata[i-2]

    if githubURLS[i-2].strip() == "no feedstock exists":
        cellref = ws.cell(row=i, column=6)
        cellref.value = githubURLS[i-2]
    else:
        cellref = ws.cell(row=i, column=6).hyperlink = githubURLS[i-2].strip()

    if anacondaURLS[i-2].strip() == "no anaconda package exists":
        cellref = ws.cell(row=i, column=7)
        cellref.value = anacondaURLS[i-2]
    else:
        cellref = ws.cell(row=i, column=7).hyperlink = anacondaURLS[i-2].strip()

for cell in ws["B"]: #highlights the values not found
   if 'not found' in cell.value:
        for cell in ws[cell.row]:
            cell.fill = redFill

wb.save("nsls_package_list.xlsx")
